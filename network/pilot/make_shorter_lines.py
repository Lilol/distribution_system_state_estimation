# -*- coding: utf-8 -*-
"""
Created on Sun Sep 12 19:26:33 2021

@author: Gergő
"""

import os

import openpyxl

import utility.configuration as configuration


# returns the only sheet of the Excel file which is specified by the argument.
def get_lines_sheet(sztring, path):
    # returns the files (Excels) from the folder of path
    workbooks_in_folder = os.listdir(path)
    # only one element is used: defined by sztring
    line_wb_name = [wb for wb in workbooks_in_folder if sztring in wb][0]
    # the load_workbook function's "load" means "Loading data..."
    line_wb = openpyxl.load_workbook(f"{path}\\{line_wb_name}")
    # only one sheet element exists
    line_sheet_name = line_wb.sheetnames[0]
    # the required sheet
    line_sheet = line_wb[line_sheet_name]
    return line_sheet


def calculate_length_ratios(coordinates):
    # Tuples represents coordinates of points.
    x_y_tuples = list(zip(coordinates[0], coordinates[1]))
    # The lengths are calculated for every part with Pithagoras Thesis
    lengths = []
    for i in range(len(x_y_tuples) - 1):
        x_0 = x_y_tuples[i][0]
        y_0 = x_y_tuples[i][1]
        x_1 = x_y_tuples[i + 1][0]
        y_1 = x_y_tuples[i + 1][1]
        lengths.append(((x_0 - x_1) ** 2 + (y_0 - y_1) ** 2) ** 0.5)
    sum_of_legths = sum(lengths)
    normed_lengths = [nl / sum_of_legths for nl in lengths]
    # return the normed lengths with the whole length (these are the ratios)
    return normed_lengths


def give_numbers_back(first, second, n):
    # get 2 points
    # interpolates n-1 point between them with equal distance
    # give back the interpolated points in order, in list
    numbers = []
    distance = second - first
    part = distance / n
    for i in range(1, n):
        numbers.append(first + part * i)
    return numbers


def insert_coordinates(indexes_numbers, list_to_insert):
    # inserts in the "list_to_insert" the interpolated coordinates
    # that interpolation creates "numbers" of lines from "indexes_numbers" to the "indexes"
    # NOTE THAT: "numbers" of lines + 1 == number of coordinates
    # e.g. indexes_numbers == [(0,2), (1,3)], list_to_insert == [10, 2, 5] ----> list_to_insert == [10, 6, 2, 3, 4, 5]
    inserting_lists = []
    for i, n in indexes_numbers:
        inserting = give_numbers_back(list_to_insert[i], list_to_insert[i + 1], n)
        inserting_lists.append([i, inserting])
    for i in range(len(inserting_lists) - 1, -1, -1):
        inserting = inserting_lists[i][1]
        where = inserting_lists[i][0]
        list_to_insert = list_to_insert[:where + 1] + inserting + list_to_insert[where + 1:]
    return list_to_insert


def index_of_too_long__number_of_parts_to_cut(list_of_full_line_lengths):
    longest_single_line = configuration.config.getfloat("network", "longest_acceptable_single_line")
    normal_length = configuration.config.getfloat("network", "normal_line_length")
    # If the line longer than longest_single_line it will cut to length//normal_length or
    # length//normal_length + 1 pieces
    # It will choose the one that results are closer to normal_length meter
    # The function gives back the index which line should be cut and the number to cut
    # If all the lengths are correct, it will be an empty list
    indexes_numbers = []
    for i, l in enumerate(list_of_full_line_lengths):
        if l > longest_single_line:
            n = int(l // normal_length)
            shorter = l / (n + 1)
            longer = l / n
            if abs(shorter - normal_length) < abs(longer - normal_length):
                indexes_numbers.append((i, n + 1))
            else:
                indexes_numbers.append((i, n))
    return indexes_numbers


def main(lines_of_the_network, path):
    line_sheet = get_lines_sheet(f"Vezetek_{path[-5:]}", path)
    # empty dict for cut lines
    new_lines_of_the_network = {}
    # iterate on every key
    for id_key in lines_of_the_network:
        # {2120911123: [[544430.69435, 544404.86835, 544377.17735, 544369.50685],
        # [168482.1442, 168476.3667, 168452.8057, 168427.1692]], 2120909339: [[544455.17085,...
        coordinates = lines_of_the_network[id_key]
        # length ratios between the coordinates
        normed_lengths = calculate_length_ratios(coordinates)
        fid_in_sheet = 0
        # skip the header
        row = 3
        while fid_in_sheet != id_key:
            # search the proper fid
            fid_in_sheet = line_sheet[f"A{str(row)}"].value
            row += 1
        row -= 1
        # When the row is found, get the length of the line
        length = line_sheet[f"D{str(row)}"].value
        # mutiply all the ratios with the full length, so we can get the real lengths
        list_of_full_line_lengths = [normed_length * length for normed_length in normed_lengths]
        # give back the indexes of line parts that are too long and give back the number of lines
        # that shoud be inserted to their place e.g. [(0, 3), (1, 7)]
        indexes_numbers = index_of_too_long__number_of_parts_to_cut(list_of_full_line_lengths)
        X_s = coordinates[0]
        Y_s = coordinates[1]
        # it inserts the new calculated coordinates to the list, where the index shows to insert
        X_s = insert_coordinates(indexes_numbers, X_s)
        Y_s = insert_coordinates(indexes_numbers, Y_s)
        # fill dict
        new_lines_of_the_network[id_key] = [X_s, Y_s]
    return new_lines_of_the_network
