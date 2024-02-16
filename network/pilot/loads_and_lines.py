# -*- coding: utf-8 -*-
"""
Created on Wed Jun  2 09:24:52 2021

@author: Gergő
"""

import os

import openpyxl
import pandas

import utility.configuration as configuration
from network.pilot import make_shorter_lines


# returns the only sheet of the Excel file which is specified by the argument.
def get_lines_sheet(sztring, path):
    # returns the files (Excels) from the folder of path
    workbooks_in_folder = os.listdir(path)
    # only one element is used: defined by sztring
    line_wb_name = [wb for wb in workbooks_in_folder if sztring in wb][0]
    # the load_workbook function's "load" means "Loading data..."
    line_wb = openpyxl.load_workbook(os.path.join(path, line_wb_name))
    # only one sheet element exists
    line_sheet_name = line_wb.sheetnames[0]
    # the required sheet
    line_sheet = line_wb[line_sheet_name]
    return line_sheet


def get_seconder_voltage_ratio(path):
    # For example: secunder voltage = 440V
    # In that case ratio will be 440V/400V = 1.1
    transformer_sheet = get_lines_sheet(f"Tr_{path[-5:]}", path)
    return float(transformer_sheet['I3'].value) / 0.4


# This function is for HMKE and vezerelet_fogyaszto objects. It returns '1' if a consumer is HMKE or vezerelet_fogyaszto
# This decision is made by its SAP datasheet (Q and R column)
def binary(value):
    if type(value) == str:
        try:
            value = float(value)
        except ValueError:
            return None
    if not value:
        return '0'
    elif value > 0.000001:
        return '1'
    else:
        return '0'


def nominal_(nappali_nevleges, rendelkezesre_allo, fazis, path):
    # It can happen that the transformers seconder voltage is not 400V. In that case we correct the 400V calculation
    # with a ratio.
    ratio = get_seconder_voltage_ratio(path)
    return_nominal = []
    # Compute the nominal power for every consumer
    for n, r, p in list(zip(nappali_nevleges, rendelkezesre_allo, fazis)):
        # 3 phase consumer
        if p == 3:
            constant_and_voltage = 3 ** 0.5 * 400 * ratio / 3  # division with 3 because of the specification of the panda modell
        # 1 phase consumer
        else:
            constant_and_voltage = 400 / (
                    3 ** 0.5) * ratio / 3  # division with 3 because of the specification of the panda modell
        # use nappali_nevleges if it is possible, if not, use rendelkezesre_allo
        try:
            return_nominal.append(constant_and_voltage * float(n) / 10 ** 6)
        except (TypeError, ValueError):
            try:
                return_nominal.append(constant_and_voltage * float(r) / 10 ** 6)
            except (TypeError, ValueError):
                # ghost consumer at transformer, on KÖF side without data
                '''ask EON'''
                return_nominal.append(0)
    return return_nominal


# This function draws the lines of the network with breakpoints. And returns a dict: {2120911123: [[
# 544430.69435, 544404.86835, 544377.17735, 544369.50685], [168482.1442, 168476.3667, 168452.8057, 168427.1692]],
# 2120909339: [[544455.17085,... The keys are the fids of the line, the values are [[...x coordinates...],
# [...y coordinates...]]
def draw_lines(line_sheet, path):
    # For first create dict
    segments_in_order = {}
    row = 2
    for fid_cell in line_sheet['A'][2:]:
        fid = fid_cell.value
        if fid in segments_in_order:
            segments_in_order[fid][0].append(line_sheet['J'][row].value)
            segments_in_order[fid][1].append(line_sheet['K'][row].value)
        else:
            segments_in_order[fid] = [[line_sheet['J'][row].value], [line_sheet['K'][row].value]]
        row += 1
    dict_for_return = {}
    # Than plot and return
    for fid in segments_in_order:
        x = segments_in_order[fid][0]
        y = segments_in_order[fid][1]
        dict_for_return[fid] = [x, y]
    # if we would like to use a bit better model of the network, we can use the model that has shorter lines
    if configuration.config.getboolean("network", "cut_to_pieces"):
        dict_for_return = make_shorter_lines.main(dict_for_return, path)
    return dict_for_return


# This function works recursively. Gets an object, its type and sheets. And returns the connected 'Ele-Vezeték
# szakasz-KIF' to the object
def find_recursively(fids, contact_types, node_sheet, assistant_line_sheet, device_sheet):
    zipped = zip(fids, contact_types)
    # fids and contact_types for next recursion
    next_fids = []
    next_contact_types = []
    for fid, contact_type in zipped:
        # Looping on objects fid, and type. For every object we open the correct sheet, defined by its type. For
        # 'Ele-Gyűjtősín-KIF': node_sheet, for 'Ele-Segédvezeték-KIF': assistant_line_sheet, for 'Ele-Készülék
        # hely-KIF': device_sheet Then we are searching for new objects (types and fids) If we find
        # 'Ele-Gyűjtősín-KIF', 'Ele-Segédvezeték-KIF' or 'Ele-Készülék hely-KIF': give them to list, call recursion (
        # last row) If we find 'Ele-Vezeték szakasz-KIF': we finished: return its fid
        if contact_type == 'Ele-Gyűjtősín-KIF':
            row_list = [i.value for i in node_sheet['A']]
            rows = [i for i, x in enumerate(row_list) if x == fid]
            new_contact_types = [node_sheet[f"G{str(row + 1)}"].value for row in rows]
            new_fids = [node_sheet[f"H{str(row + 1)}"].value for row in rows]
        elif contact_type == 'Ele-Segédvezeték-KIF':
            row_list = [i.value for i in assistant_line_sheet['A']]
            rows = [i for i, x in enumerate(row_list) if x == fid]
            new_contact_types = [assistant_line_sheet[f"G{str(row + 1)}"].value for row in rows]
            new_fids = [assistant_line_sheet[f"H{str(row + 1)}"].value for row in rows]
        elif contact_type == 'Ele-Készülék hely-KIF':
            row_list = [i.value for i in device_sheet['A']]
            rows = [i for i, x in enumerate(row_list) if x == fid]
            new_contact_types = [device_sheet[f"J{str(row + 1)}"].value for row in rows]
            new_fids = [device_sheet[f"K{str(row + 1)}"].value for row in rows]
        elif contact_type == 'Ele-Vezeték szakasz-KIF':
            return fid
        else:
            raise RuntimeError(f"Unknown contact type: {contact_type}")
        next_contact_types += new_contact_types
        next_fids += new_fids
    return find_recursively(next_fids, next_contact_types, node_sheet, assistant_line_sheet, device_sheet)


# This function gets the contact_types, contacts_fids and returns the type of 'Ele-Gyűjtősín-KIF' contacts main_contacts
# These main contacts are the fids of the nearest (connected) 'Ele-Vezeték szakasz-KIF's
def get_nodes_places_as_dict(contact_types, contacts_fids, path):
    # list for 'Ele-Gyűjtősín-KIF's
    list_of_nodes = []
    # fill the list
    for contact_type, contacts_fid in zip(contact_types, contacts_fids):
        if contact_type == 'Ele-Gyűjtősín-KIF' and contacts_fid not in list_of_nodes:
            list_of_nodes.append(contacts_fid)
    # get KIF-modell-Gys_12345.xlsx
    node_sheet = get_lines_sheet(f"Gys_{path[-5:]}", path)
    # get KIF-modell-Segedvez_12345.xlsx
    assistant_line_sheet = get_lines_sheet(f"Segedvez_{path[-5:]}", path)
    # get KIF-modell-Kesz_12345.xlsx
    device_sheet = get_lines_sheet(f"Kesz_{path[-5:]}", path)
    # We are searching (recursively) main_contacts for 'Ele-Gyűjtősín-KIF' fids
    contact_types_in_func = ['Ele-Gyűjtősín-KIF']
    # main contacts in order, check the  'find_recursively' functions comments above
    main_contacts = [find_recursively([fid], contact_types_in_func, node_sheet, assistant_line_sheet, device_sheet) for
                     fid in list_of_nodes]
    # make and fill a dict, where keys are the'Ele-Gyűjtősín-KIF's fids and values are 'Ele-Vezeték szakasz-KIF's fids
    dict_of_nodes_with_main_contacts = {}
    for n, m in zip(list_of_nodes, main_contacts):
        dict_of_nodes_with_main_contacts[n] = m
    return dict_of_nodes_with_main_contacts


def get_loads_of_network(path):
    workbooks_in_folder = os.listdir(path)
    line_wb_name = \
        [wb for wb in workbooks_in_folder if
         f'Fogyasztok_{configuration.config.get("input", "network_id").value}' in wb][
            0]
    path_to_sheet = os.path.join(path, line_wb_name)
    load_info = pandas.read_excel(path_to_sheet, engine="openpyxl", sheet_name=0)

    # in column A the fids can be found
    fids_of_loads = load_info["FID"].astype(int).values
    # in column K you can find 'Ele-Vezeték szakasz-KIF' or 'Ele-Gyűjtősín-KIF'
    contact_types = load_info["Kapcsolódó Objektum név"].values
    # in column L you can find 'Ele-Vezeték szakasz-KIF's or 'Ele-Gyűjtősín-KIF's fid
    contacts_fids = load_info["Kapcsolódó Objketum példány"].astype(int).values
    # If the contact_type is 'Ele-Gyűjtősín-KIF', we have to find the fid of the nearest (connected) 'Ele-Vezeték
    # szakasz-KIF's fid These are the main_contacts
    # in column M the EOV X can be found
    x_coordinates = load_info["EOV X"].values
    # in column N the EOV Y can be found
    y_coordinates = load_info["EOV Y"].values
    try:
        # from column Q we can get if there is a HMKE or not
        hmke_0_or_1 = load_info[load_info.columns.str.contains("hmke")[0]].astype("bool").values
        # from column R we can get if there is a vezerelt_fogyaszto or not
        vezerelt_fogy_0_or_1 = load_info[load_info.columns.str.contains("vezerelt")[0]].astype("bool").values
        # from column S we can get nappali_nevleges power
        nappali_nevleges = load_info[load_info.columns.str.contains("nappali")[0]].astype("bool").values
        # from column W we can get if there is 1 or 3 phase
        fazis = load_info[load_info.columns.str.contains("fazis")[0]].astype("bool").values
        # from column T we can get rendelkezesre_allo power
        rendelkezesre_allo = load_info["rendelkezesre_allo"].astype("bool").values
        # calculate the nominal power, depends on: nappali_nevleges, rendelkezesre_allo, fazis
        nominal = nominal_(nappali_nevleges, rendelkezesre_allo, fazis, path)
    except Exception as e:
        print(f"Sheet not found '{e}'")
        nominal = [0] * len(y_coordinates)
        hmke_0_or_1 = nominal
        vezerelt_fogy_0_or_1 = nominal
        nappali_nevleges = nominal
        rendelkezesre_allo = nominal
        fazis = nominal
    return fids_of_loads, contact_types, contacts_fids, x_coordinates, y_coordinates, hmke_0_or_1, vezerelt_fogy_0_or_1, \
        nappali_nevleges, rendelkezesre_allo, fazis, nominal


def create_pilot_network(path):
    fids_of_loads, contact_types, contacts_fids, x_coordinates, y_coordinates, hmke_0_or_1, vezerelt_fogy_0_or_1, \
        nappali_nevleges, rendelkezesre_allo, fazis, nominal = get_loads_of_network(path)
    if 'Ele-Gyűjtősín-KIF' in contact_types:
        # If the type is 'Ele-Gyűjtősín-KIF': we change its fid in contacts_fids to its main contacts fid
        nodes_places_as_dict = get_nodes_places_as_dict(contact_types, contacts_fids, path)
        for i, contact_type in enumerate(contact_types):
            if contact_type == 'Ele-Gyűjtősín-KIF':
                contacts_fids[i] = nodes_places_as_dict[contacts_fids[i]]

    # create and plot the loads of the network
    loads_of_the_network = [load for load in
                            zip(fids_of_loads, contact_types, contacts_fids, x_coordinates, y_coordinates, nominal,
                                hmke_0_or_1, vezerelt_fogy_0_or_1) if 'Transzfor' not in str(load[1])]

    # create and plot the lines of the network
    line_sheet = get_lines_sheet('Vezetek_graf', path)
    return draw_lines(line_sheet, path), loads_of_the_network
