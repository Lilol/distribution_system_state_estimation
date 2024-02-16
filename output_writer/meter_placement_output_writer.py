from logging import getLogger
from os.path import exists

from openpyxl import load_workbook, Workbook
from openpyxl.styles import Font

import utility.configuration as configuration
from output_writer.assemble_path import assemble_file
from output_writer.output_writer import OutputWriter
from smart_meter_placement.accuracy_statistics import accuracy_statistics


class MeterPlacementOutputWriter(OutputWriter):
    def __init__(self):
        super(MeterPlacementOutputWriter, self).__init__()
        self._filename = configuration.config.get("output", "meter_placement_output_file", fallback="meter_placement")
        self._logger = getLogger("output_writer.meter_placement")
        self.__meters_dumped = len(accuracy_statistics.get_meters_fixed_sofar())
        self.__reinit_necessary = True
        self.__writer = None

    def save(self):
        pass

    def initialize(self, multi_unit_run=False, **kwargs):
        if not configuration.config.getboolean("scenario", "meter_positioning"):
            return

        if self.__reinit_necessary:
            self.__reinit_necessary = False
            self.__meters_dumped = len(accuracy_statistics.get_meters_fixed_sofar())
            self._output_path = assemble_file(self._filename, self._mode, appendages=False)
            self._logger.info(f"Initializing new output file: '{self._output_path}'")

    def write_step(self):
        pass

    def write_final(self):
        if not configuration.config.getboolean("scenario", "meter_positioning"):
            return

        metered_bus = configuration.config.getint("scenario", "meter_positioning_bus")

        if not exists(self._output_path):
            wb = Workbook()
            wb.save(self._output_path)

        wb = load_workbook(self._output_path)
        sheet = f"circuit_{self.__meters_dumped + 1}"
        if sheet not in wb.sheetnames:
            wb.create_sheet(sheet)
            ws = wb[sheet]
            ws['A1'] = "bus"
            ws['B1'] = "precision"
            ws['A1'].font = Font(bold=True)
            ws['B1'].font = Font(bold=True)

        ws = wb[sheet]
        n = accuracy_statistics.n_finished_buses() + 1
        ws[f"A{str(n)}"] = metered_bus
        ws[f"B{str(n)}"] = accuracy_statistics.get_accuracy(metered_bus)
        wb.save(self._output_path)

        fixed_meters = accuracy_statistics.get_meters_fixed_sofar()
        currently_metered_circuit = len(fixed_meters)

        if 0 < currently_metered_circuit and self.__meters_dumped < currently_metered_circuit:
            self.__meters_dumped += 1
            letter = chr(self.__meters_dumped + 65)
            circuits = accuracy_statistics.get_ordered_circuits()
            circuit = circuits[currently_metered_circuit - 1 - accuracy_statistics.get_first_non_empty_circuit_num()]

            sheet = "optimal_meter_placement"
            if sheet not in wb.sheetnames:
                wb.create_sheet(sheet)
                ws = wb[sheet]
                ws['A1'] = "Circuit no."
                ws['A2'] = "Yearly consumption (kWh)"
                ws['A3'] = "Optimal meter placement bus"
                ws['A4'] = "Precision"
                ws['A1'].font = Font(bold=True)
                ws['A2'].font = Font(bold=True)
                ws['A3'].font = Font(bold=True)
                ws['A4'].font = Font(bold=True)

            ws = wb[sheet]
            ws[f"{letter}1"] = currently_metered_circuit
            ws[f"{letter}2"] = circuit.sum_yc
            ws[f"{letter}3"] = fixed_meters[-1]
            ws[f"{letter}4"] = accuracy_statistics.get_fixed_meter_precisions()[-1]
            wb.save(self._output_path)

            if self.__meters_dumped == accuracy_statistics.n_circuits():
                self.__reinit_necessary = True


class MeterPlacementOutputWriterPvScenario(MeterPlacementOutputWriter):
    def write_final(self):
        if not configuration.config.getboolean("scenario", "meter_positioning"):
            return

        metered_bus = configuration.config.getint("scenario", "meter_positioning_bus")

        if not exists(self._output_path):
            wb = Workbook()
            wb.save(self._output_path)

        wb = load_workbook(self._output_path)
        sheet = "results"
        if sheet not in wb.sheetnames:
            wb.create_sheet(sheet)
            ws = wb[sheet]
            ws['A1'] = "bus"
            ws['B1'] = "precision"
            ws['A1'].font = Font(bold=True)
            ws['B1'].font = Font(bold=True)

        ws = wb[sheet]
        n = accuracy_statistics.n_finished_buses() + 1
        ws[f"A{str(n)}"] = metered_bus
        ws[f"B{str(n)}"] = accuracy_statistics.get_accuracy(metered_bus)
        wb.save(self._output_path)

        if 0 < accuracy_statistics.any_meter_fixed():
            pv_circuit = accuracy_statistics.get_pv_circuit()

            sheet = "optimal_meter_placement"
            if sheet not in wb.sheetnames:
                wb.create_sheet(sheet)
                ws = wb[sheet]
                ws['A1'] = "Circuit no."
                ws['A2'] = "Yearly consumption (kWh)"
                ws['A3'] = "Sum annual PV power (kWh)"
                ws['A4'] = "Optimal meter placement bus"
                ws['A5'] = "Precision"
                ws['A1'].font = Font(bold=True)
                ws['A2'].font = Font(bold=True)
                ws['A3'].font = Font(bold=True)
                ws['A4'].font = Font(bold=True)
                ws['A5'].font = Font(bold=True)

            ws = wb[sheet]
            fixed_meters = accuracy_statistics.get_meters_fixed_sofar()
            ws['B1'] = accuracy_statistics.get_pv_circuit_index()
            ws['B2'] = pv_circuit.sum_yc
            ws['B3'] = pv_circuit.sum_pv_kwp * 1200
            ws['B4'] = fixed_meters[-1]
            ws['B5'] = accuracy_statistics.get_fixed_meter_precisions()[-1]
            wb.save(self._output_path)

            self.__reinit_necessary = True
