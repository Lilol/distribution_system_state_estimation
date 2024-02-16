from logging import getLogger
from os.path import join

from openpyxl import load_workbook
from pandas import read_excel

import utility.configuration as configuration
from network.definitions import NetworkKind


def get_fixed_meters(circuit_idx=0, from_config=False):
    if from_config:
        return [int(s) for s in configuration.config.get("scenario", "fixed_meters")]

    fixed_meters = dict()
    if circuit_idx == 0:
        return fixed_meters

    try:
        input_frame = read_excel(join(configuration.config.get("paths", "output_path"),
                                      NetworkKind.EXACT.value,
                                      f"meter_placement_sc{str(configuration.config.get('scenario', 'scenario').value)}.xlsx"),
                                 sheet_name="optimal_meter_placement", engine='openpyxl', header=[0], index_col=0)
    except KeyError as e:
        getLogger("smart_meter_placement.checkpoint_restart").warning(f"Worksheet not found: {str(e)}")
        raise e
    else:
        fixed_meters = dict(zip(input_frame.loc["Optimal meter placement bus"].head(circuit_idx).to_numpy()
                                .astype('int'), input_frame.loc["Precision"].head(circuit_idx).to_numpy()))
    finally:
        return fixed_meters


def get_accuracies(circuit_idx):
    if configuration.config.has_option("scenario", "restart_from_bus"):
        start_from = configuration.config.getint("scenario", "restart_from_bus")
    else:
        getLogger("smart_meter_placement.checkpoint_restart").warning("Bus to restart from must be specified when "
                                                                      "checkpoint restart is active!")
        return []

    wb = load_workbook(join(configuration.config.get("paths", "output_path"),
                            NetworkKind.EXACT.value, "meter_placement_sc" +
                            configuration.config.get("scenario", "scenario").value + ".xlsx"))

    sheet = f"circuit_{str(circuit_idx + 1)}"
    if sheet not in wb.sheetnames:
        getLogger("smart_meter_placement.checkpoint_restart").warning(f"Sheet not found: {sheet}")
        return []

    accuracies = dict()
    ws = wb[sheet]
    for cell in ws.iter_rows(min_row=2):
        if cell[0].value == start_from:
            break
        accuracies[cell[0].value] = cell[1].value
    return accuracies


def get_accuracies_and_fixed_meters(circuit_idx):
    return get_fixed_meters(circuit_idx), get_accuracies(circuit_idx)
